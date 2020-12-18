# -*- coding:utf-8 -*-
#  extract_DTCList_to_json.py
#
#  ~~~~~~~~~~~~
#
#  use to extract DTC list information from excel to .json
#
#  ~~~~~~~~~~~~
#
#  ------------------------------------------------------------------
#  Author : CAO Cheng    
#  Create : 12.11.2020
#  Last change : 13.11.2020
#
#  Language: Python 3.6
#  ------------------------------------------------------------------

import json

from openpyxl import load_workbook

file = "C Platform-Diagnostic Questionnaires-ABS_ESC_20201014.xlsx" 
#file = "123.xlsx" 
un_use_str = '123.'


class ExcelUtils:
    def __init__(self):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[sheets.index('DTCList')]
        self.ws = self.wb[self.sheet]

    # 行数
    def get_rows(self):
        rows = self.ws.max_row
        return rows

    # 列数
    def get_clos(self):
        clo = self.ws.max_column
        return clo

    # 获取值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 修改值并保存
    def set_cell_value(self, row, column, cell_value):
        try:
            self.ws.cell(row=row, column=column).value = cell_value
            self.wb.save(self.file)
        except Exception as e:
            print("error :{}".format(e))
            self.wb.save(self.file)

    # 替换单元格中的内容
    def replace_cell_value(self):
        # 遍历第一行的值，
        for i in range(1, self.get_clos() + 1):
            cell_value = self.get_cell_value(1, i)
            # 是否存在需要替换的值
            if un_use_str in cell_value:
                cell_replace = cell_value.replace(un_use_str, "")
                self.set_cell_value(1, i, cell_replace)


def to_json(filename):
    excel_utils = ExcelUtils()
    excel_dict = {}

    # 遍历excel中的值存入字典中
    for i in range(8, 209):
        dict_key_1 = int(excel_utils.get_cell_value(i, 3),16)
        dict_key_2 = int(excel_utils.get_cell_value(i, 5),16)
        dict_key = (dict_key_1 << 8) | dict_key_2
        dict_value = excel_utils.get_cell_value(i, 12).replace("\n","")
        print(dict_value)
        excel_dict[hex(dict_key)] = dict_value
    # 字典转json
    with open(filename,'w') as file_obj:
        json.dump(excel_dict,file_obj, ensure_ascii=False)
    return

if __name__ == "__main__":
    filename = 'DTCList.json'
    to_json(filename)
